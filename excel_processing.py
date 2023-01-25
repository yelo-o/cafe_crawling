# 모듈(엑셀 후처리)
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from crawling import *

# 외부 변수
# from crawling import crawling_time
# from crawling import srh_krd

# 엑셀 후처리
def excel_processing():
    # dir = f'C:/Users/user/Desktop/VENVWorkspace/cafe_crawling/naver cafe crawling_2023-01-16 11.00.05.xlsx' # 테스트용
    dir = f'C:/Users/user/Desktop/VENVWorkspace/cafe_crawling/{crawling_time}-{srh_krd}검색.xlsx'
    wb = load_workbook(dir) # 해당 경로의 엑셀 파일 불러오기
    ws = wb.active # 불러온 엑셀 파일의 시트 활성화
    
    # 중복 데이터 삭제

    # A열 가운데 정렬
    for cell in range(len(ws['A'])):
        ws['A'+str(cell+2)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    # C열 가운데 정렬
    # for cell in range(len(ws['C'])):
    #     ws['C'+str(cell+1)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    # E열 가운데 정렬
    for cell in range(len(ws['E'])):
        ws['E'+str(cell+1)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    ws['B1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    ws['D1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    # 열 너비 변경
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 9.7
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 35

    # 자동 줄바꿈
    for cell in range(len(ws['B'])):
        ws['B'+str(cell+2)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
    for cell in range(len(ws['C'])):
        ws['C'+str(cell+2)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
    for cell in range(len(ws['D'])):
        ws['D'+str(cell+2)].alignment = openpyxl.styles.Alignment(vertical = 'center', wrap_text = True)

    # 제목 행 색상 넣기
    y_color = PatternFill(start_color = 'ffff99', end_color = 'ffff99', fill_type = 'solid')
    ws.cell(1,1).fill = y_color
    ws.cell(1,2).fill = y_color
    ws.cell(1,3).fill = y_color
    ws.cell(1,4).fill = y_color
    ws.cell(1,5).fill = y_color

    # 저장 및 닫기
    # wb.save(filename = f'naver cafe crawling_2023-01-16 11.00.05.xlsx') # 테스트용
    wb.save(filename = f'{crawling_time}-{srh_krd}검색.xlsx')
    wb.close()