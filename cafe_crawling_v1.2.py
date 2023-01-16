# 크롤링 관련 모듈
import requests
import urllib
import time
from datetime import datetime, timedelta
from datetime import date
from dateutil.relativedelta import relativedelta
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import sys, os
import pyperclip
import pyautogui as pg
from bs4 import BeautifulSoup as bs
import pandas as pd
from credentials import u_id, u_pw
# 메일 관련 모듈
import smtplib
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
# 엑셀 후처리
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# 로그인
"""
pyperclip 모듈을 이용해서 아이디, 비밀번호 복사 붙여넣기
    - > 일반 셀레니움으로 입력하면 네이버에서 크롤링 봇 감지하여 자동입력방지
"""
def crawling_process():
    # 실행 절차
    global browser
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"]) # 쓸모없는 로그 삭제
    browser = webdriver.Chrome(options=options)
    browser.maximize_window() # 윈도우 최대로 확대
    url = "https://cafe.naver.com/ak573/"
    browser.get(url)

    # 로그인 화면 이동
    btnToLogin = browser.find_element(By.XPATH,'//*[@id="gnb_login_button"]')
    btnToLogin.click()
    today = str(date.today()) # 오늘 날짜 '2023-01-06' 과 같은 형식으로 지정
    login()
    keyword_search()
    browser.switch_to.frame('cafe_main') # iframe 진입하기

    # 게시글 50개씩 보이게 하기
    browser.find_element(By.CSS_SELECTOR,"#listSizeSelectDiv").click()
    browser.find_element(By.XPATH,"/html/body/div[1]/div/div[3]/div/div[3]/ul/li[7]/a").click()

    collect_url()
    moveToUrl()
    data_to_excel()
    browser.quit()
    excel_processing()
    # pg.alert(f"<{srh_krd}> 키워드로 검색한 결과 중 최근 50개의 게시글을 크롤링 완료하였습니다!")
    
def login():
    input_id = browser.find_element(By.XPATH,'//*[@id="id"]')
    input_pw = browser.find_element(By.XPATH,'//*[@id="pw"]')
    btn_lgn = browser.find_element(By.XPATH,'//*[@id="log.login"]')
    input_id.click()
    pyperclip.copy(u_id)
    input_id.send_keys(Keys.CONTROL,'v')
    input_pw.click()
    pyperclip.copy(u_pw)
    input_pw.send_keys(Keys.CONTROL,'v')
    btn_lgn.click()
    
def keyword_search():
    global srh_krd
    srh = browser.find_element(By.XPATH,'//*[@id="topLayerQueryInput"]') # 검색창
    srh_krd = '선물' # 검색 키워드 변수 저장
    srh.send_keys(srh_krd)
    btn_srh = browser.find_element(By.XPATH,'//*[@id="cafe-search"]/form/button')
    btn_srh.click()
    time.sleep(3)

def collect_url():
    # 리스트 화면에 있는 글번호 담기
    global numList, urls
    numList = [i.text for i in browser.find_elements(By.CSS_SELECTOR,'.inner_number')]
    urls = []
    url = 'https://cafe.naver.com/ak573/'
    for num in numList:
        urls.append(url+num)
        
    # 제목 담기
    global title_list
    title_list = [i.text for i in browser.find_elements(By.CSS_SELECTOR,'.article')]
    
        
# 변수 설정
content_data = [] # 데이터 담을 리스트
time_data = []
title_data = []


def collect_data():
    global content_data, time_data, title_data
    browser.switch_to.frame('cafe_main')
    # 타이틀
    try:
        written_title = browser.find_elements(By.XPATH, '//*[@id="app"]/div/div/div[2]/div[1]/div[1]/div/h3')[0]
        title_data.append(written_title.text)
    except:
        title_data.append("**접근이 불가한 게시판입니다.**")
        pass
    
    # 글 작성일자 및 시간
    try:
        written_time = browser.find_elements(By.XPATH, '//*[@id="app"]/div/div/div[2]/div[1]/div[2]/div/div[2]/span[1]')[0]
        time_data.append(written_time.text)
    except:
        time_data.append("**접근이 불가한 게시판입니다.**")
        pass
    
    # 글 내용
    try:
        written_content = browser.find_elements(By.XPATH, '//*[@id="app"]/div/div/div[2]/div[2]/div[1]/div/div[1]')[0]
        content_data.append(written_content.text)
    except:
        content_data.append("**접근이 불가한 게시판입니다.**")
        pass

def moveToUrl():
    for url in urls:
        browser.get(url)
        time.sleep(1)
        collect_data()

# 엑셀에 데이터 저장        
def data_to_excel():
    index_list = list(range(1, len(urls)+1)) # list(range(1,51)) # 1~50까지 넘버링
    total_data = pd.DataFrame({"제목" : title_list, "날짜" : time_data, "내용" : content_data, "url" : urls}, index = index_list)
    total_data.index.name = "No."
    total_data.to_excel(f"naver cafe crawling_{crawling_time}.xlsx",index=True)

# 엑셀 후처리
def excel_processing():
    dir = f'C:/Users/user/Desktop/VENVWorkspace/cafe_crawling/naver cafe crawling_{crawling_time}.xlsx'
    wb = load_workbook(dir) # 해당 경로의 엑셀 파일 불러오기
    ws = wb.active # 불러온 엑셀 파일의 시트 활성화

    # A열 가운데 정렬
    for cell in range(len(ws['A'])):
        ws['A'+str(cell+2)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    # C열 가운데 정렬
    # for cell in range(len(ws['C'])):
    #     ws['C'+str(cell+1)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    # E열 가운데 정렬
    for cell in range(len(ws['E'])):
        ws['E'+str(cell+1)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    ws['B1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    ws['D1'].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

    # 열 너비 변경
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 9.7
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 35

    # 자동 줄바꿈
    for cell in range(len(ws['B'])):
        ws['B'+str(cell+2)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
    for cell in range(len(ws['C'])):
        ws['C'+str(cell+2)].alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
    for cell in range(len(ws['D'])):
        ws['D'+str(cell+2)].alignment = openpyxl.styles.Alignment(vertical = 'center', wrap_text = True)

    # 제목 행 색상 넣기
    y_color = PatternFill(start_color = 'ffff99', end_color = 'ffff99', fill_type = 'solid')
    ws.cell(1,1).fill = y_color
    ws.cell(1,2).fill = y_color
    ws.cell(1,3).fill = y_color
    ws.cell(1,4).fill = y_color
    ws.cell(1,5).fill = y_color

    # 저장 및 닫기
    wb.save(filename = f'naver cafe crawling_{crawling_time}.xlsx')
    wb.close()

# 메일보내기 함수
def mailing():
    smtp_etners = smtplib.SMTP('mail.etners.com', 25)
    smtp_etners.ehlo()
    msg = MIMEMultipart()

    # 제목 입력
    # msg['subject'] = f'[{crawling_time}]네이버 카페 크롤링 공유의 件'
    msg['subject'] = f'[인사쟁이 카페 크롤링] {crawling_time} 件'
    # 보내는 사람
    msg['From'] = '김민규 <flyordig@etners.com>'
    # 받는 사람
    # msg['To'] = ", ".join(['hjs0131@etners.com', 'orion@etners.com']) # 수신자가 다수일 때
    # msg['To'] = 'hjs0131@etners.com' # 한진솔 프로
    msg['To'] = 'flyordig@etners.com' # 테스트용(본인 계정)
    # 참조
    # msg['Cc'] = ", ".join(['orion@etners.com', 'flyordig@etners.com'])
    # 내용 입력
    body = MIMEText(f" 안녕하세요,\n\n 경영혁신그룹 크롤링 봇입니다. \n\n {crawling_time} 에 [인사쟁이 카페]에서 '선물' 키워드로 크롤링 진행한 엑셀 파일 전달드립니다.\n\n감사합니다.", _charset = 'utf-8')
    msg.attach(body)

    # 첨부 파일
    path = f'C:/Users/user/Desktop/VENVWorkspace/cafe_crawling/naver cafe crawling_{crawling_time}.xlsx'
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(open(path, "rb").read())
    encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(path))
    msg.attach(part)

    # 메일 송신
    smtp_etners.send_message(msg)

def crawlNmail():
    global now, st_now, crawling_time
    now = datetime.now() # 시간 변수
    st_now = str(now)
    crawling_time = st_now[0:19].replace(':',".")
    print(crawling_time)
    crawling_process()
    mailing()

crawlNmail()
    